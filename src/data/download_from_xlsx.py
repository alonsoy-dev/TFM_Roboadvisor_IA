import os, re, time, argparse
from urllib.parse import urljoin, urlparse
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from datetime import datetime, timedelta
from excel_convert import convert_with_excel

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/128 Safari/537.36")
HDR = {"User-Agent": UA, "Accept": "*/*"}

def sanitize(s): return re.sub(r"[^A-Za-z0-9._-]+", "_", s).strip("_")
def ensure_dir(p): os.makedirs(p, exist_ok=True)

def _norm_header(s: str) -> str:
    s = str(s or "").strip().lower()
    return re.sub(r"[.\s_-]+", "", s)

def find_header(ws):
    # Busca la fila de cabecera que contenga 'Ticker'
    for r in range(1, min(ws.max_row, 30)+1):
        vals = [(ws.cell(r,c).value or "") for c in range(1, ws.max_column+1)]
        norm = [_norm_header(v) for v in vals]
        if "ticker" in norm:
            return r
    return None

def parse_incept_date(raw):
    if raw is None or str(raw).strip() == "":
        return None
    if hasattr(raw, "date"):
        try:
            return raw.date()
        except Exception:
            pass
    if isinstance(raw, (int, float)):
        base = datetime(1899, 12, 30)
        try:
            return (base + timedelta(days=int(raw))).date()
        except Exception:
            pass
    s = str(raw).strip()
    fmts = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d-%b-%Y", "%b %d, %Y"]
    for f in fmts:
        try:
            return datetime.strptime(s, f).date()
        except Exception:
            continue
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None

def guess_ext(resp):
    ct = (resp.headers.get("Content-Type") or "").lower()
    path = urlparse(resp.url).path.lower()
    if "text/csv" in ct or path.endswith(".csv"):       return ".csv"
    if "spreadsheetml" in ct or path.endswith(".xlsx"): return ".xlsx"
    if "ms-excel" in ct or path.endswith(".xls"):       return ".xls"
    if path.endswith(".zip"):                           return ".zip"
    return ".dat"

def iter_rows_with_links(xlsx_path, cutoff_date=None, verbose=False):
    wb = load_workbook(xlsx_path, data_only=True, read_only=False)
    total = ok = 0
    for ws in wb.worksheets:
        header_row = find_header(ws)
        if not header_row:
            if verbose: print(f"[WARN] Hoja '{ws.title}': no se encontró cabecera 'Ticker'")
            continue

        headers = [(ws.cell(header_row,c).value or "") for c in range(1, ws.max_column+1)]
        norm = [_norm_header(h) for h in headers]

        def find_col(candidates):
            for k in candidates:
                if k in norm: return norm.index(k)+1
            return None

        col_ticker = find_col(["ticker"])
        if not col_ticker:
            if verbose: print(f"[WARN] Hoja '{ws.title}': columna Ticker no encontrada")
            continue
        col_name   = find_col(["name"])
        col_incept = find_col(["inceptdate","inceptiondate","inception","incept"])

        r = header_row+1
        while r <= ws.max_row:
            total += 1
            tcell = ws.cell(r, col_ticker)
            ticker = str(tcell.value or "").strip()
            if not ticker:
                r += 1; continue

            # Filtro por fecha
            if cutoff_date and col_incept:
                inception_raw = ws.cell(r, col_incept).value
                inception = parse_incept_date(inception_raw)
                if inception is None:
                    if verbose: print(f"[SKIP] {ticker}: Incept.Date inválida ({inception_raw!r})")
                    r += 1; continue
                if inception > cutoff_date:
                    if verbose: print(f"[SKIP] {ticker}: {inception} > {cutoff_date}")
                    r += 1; continue
            elif cutoff_date and not col_incept:
                if verbose: print(f"[WARN] Hoja '{ws.title}': sin columna Incept.Date; no se aplica filtro a esta hoja")

            # Localiza la URL
            url = None
            if tcell.hyperlink and tcell.hyperlink.target:
                url = tcell.hyperlink.target
            elif isinstance(tcell.value, str) and tcell.value.upper().startswith("=HYPERLINK("):
                m = re.search(r'"(https?://[^"]+)"', tcell.value)
                if m: url = m.group(1)
            if not url:
                for c in range(1, ws.max_column+1):
                    cell = ws.cell(r, c)
                    if cell.hyperlink and cell.hyperlink.target:
                        url = cell.hyperlink.target; break
                    if isinstance(cell.value, str) and cell.value.upper().startswith("=HYPERLINK("):
                        m = re.search(r'"(https?://[^"]+)"', cell.value)
                        if m: url = m.group(1); break
            if url:
                name = str(ws.cell(r, col_name).value).strip() if col_name else ""
                ok += 1
                if verbose: print(f"[OK] {ticker} → {url}")
                yield {"ticker": ticker, "name": name, "url": url}
            else:
                if verbose: print(f"[SKIP] Sin URL en fila {r}")
            r += 1

    if verbose:
        print(f"[RESUMEN] {ok}/{total} filas con URL tras filtro")

def find_download_link(session, product_url):
    r = session.get(product_url, headers=HDR, timeout=30)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "lxml")
    keyword = "data download"
    # Por texto
    for a in soup.find_all("a"):
        txt = (a.get_text() or "").strip().lower()
        if any(k in txt for k in keyword):
            href = a.get("href")
            if href: return urljoin(r.url, href)
    # Por patrón en href
    for a in soup.find_all("a", href=True):
        href = a["href"].lower()
        if any(k in href for k in ("download","filetype=","ajax","csv","xlsx")):
            return urljoin(r.url, a["href"])
    return None

def download_file(session, url, out_stub):
    with session.get(url, headers=HDR, timeout=60, stream=True) as resp:
        resp.raise_for_status()
        ext = guess_ext(resp)
        out = os.path.splitext(out_stub)[0] + ext
        with open(out, "wb") as f:
            for chunk in resp.iter_content(1<<14):
                if chunk: f.write(chunk)
    return out

def main():
    ap = argparse.ArgumentParser(description="Descarga 'Data Download' de iShares desde un .xlsx (hipervínculos en Ticker).")
    ap.add_argument("xlsx", help="Ruta al .xlsx")
    ap.add_argument("--out", default="ishares_downloads", help="Carpeta de salida")
    ap.add_argument("--sleep", type=float, default=1.0, help="Pausa entre descargas (s)")
    ap.add_argument("--max", type=int, default=None, help="Procesar solo los primeros N (opcional)")
    ap.add_argument("--date", type=str, help="Filtra Incept.Date <= esta fecha (YYYY-MM-DD)")
    ap.add_argument("--zip", action="store_true", help="Comprimir ficheros descargados")
    ap.add_argument("--verbose", action="store_true", help="Log detallado")
    args = ap.parse_args()

    cutoff = None
    if args.date:
        cutoff = datetime.strptime(args.date, "%Y-%m-%d").date()

    ensure_dir(args.out)
    session = requests.Session(); session.headers.update(HDR)

    rows = list(iter_rows_with_links(args.xlsx, cutoff_date=cutoff, verbose=args.verbose))
    if args.verbose:
        print(f"[INFO] Filas con URL detectadas (tras filtro): {len(rows)}")
    if args.max: rows = rows[:args.max]

    downloaded_paths = []
    for rec in rows:
        display = rec["name"] or rec["ticker"]
        base = sanitize(rec["ticker"])
        try:
            dl = find_download_link(session, rec["url"])
            if not dl:
                print(f'Error en la descarga de "{display}"')
                time.sleep(args.sleep); continue

            path = download_file(session, dl, os.path.join(args.out, base))
            if path.lower().endswith(".xls"):
                try:
                    xlsx_path = convert_with_excel(path)
                    try:
                        os.remove(path)
                    except Exception:
                        pass
                    path = xlsx_path
                except Exception as e:
                    print(
                        f'[WARN] Conversión Excel fallida para "{os.path.basename(path)}": {e}. Se deja el .xls original.')

            downloaded_paths.append(path)

            print(f'"{display}" descargado con éxito')
            time.sleep(args.sleep)
        except Exception:
            print(f'Error en la descarga de "{display}"')

    if args.zip and downloaded_paths:
        zip_name = os.path.splitext(os.path.basename(args.xlsx))[0] + ".zip"
        zip_path = os.path.join(args.out, zip_name)
        import zipfile
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for p in downloaded_paths:
                # Incluye formatos típicos de datos
                if p.lower().endswith((".xlsx", ".xls", ".csv", ".zip")):
                    zf.write(p, arcname=os.path.basename(p))
                    os.remove(p)

if __name__ == "__main__":
    print("Uso TEST: python script.py  <ruta.xlsx> --out <ruta:out> --sleep 1.0 --max 3 --zip --verbose")
    print("Uso: python script.py  <ruta.xlsx> --out <ruta:out> --sleep 1.0 --date 2010-09-01 --zip --verbose")
    main()